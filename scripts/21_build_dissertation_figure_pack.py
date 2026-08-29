from __future__ import annotations

import json
import os
import shutil
from pathlib import Path

import numpy as np
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
MPL_CONFIG = ROOT / ".mplconfig"
MPL_CONFIG.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(MPL_CONFIG))

import matplotlib  # noqa: E402

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.lines import Line2D  # noqa: E402
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


OUT = ROOT / "results" / "dissertation_figure_table_pack"
FIG_DIR = OUT / "figures"
TABLE_DIR = OUT / "tables"
SOURCE_DIR = OUT / "source_data"
for directory in (OUT, FIG_DIR, TABLE_DIR, SOURCE_DIR):
    directory.mkdir(parents=True, exist_ok=True)


COLORS = {
    "blue": "#3B6F8E",
    "blue_light": "#B9CEDA",
    "orange": "#C8743F",
    "gold": "#D3A33F",
    "green": "#2E7D66",
    "red": "#B3403E",
    "purple": "#79689A",
    "gray": "#AEB7BC",
    "gray_light": "#E5E9EB",
    "dark": "#27343A",
}

matplotlib.rcParams.update(
    {
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "DejaVu Sans", "Liberation Sans"],
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "font.size": 8.5,
        "axes.titlesize": 10,
        "axes.labelsize": 9,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "legend.fontsize": 7.5,
        "axes.spines.right": False,
        "axes.spines.top": False,
        "axes.linewidth": 0.8,
        "legend.frameon": False,
        "figure.facecolor": "white",
    }
)


def save_figure(fig: plt.Figure, stem: str) -> None:
    fig.savefig(FIG_DIR / f"{stem}.svg", bbox_inches="tight", facecolor="white")
    fig.savefig(FIG_DIR / f"{stem}.pdf", bbox_inches="tight", facecolor="white")
    fig.savefig(FIG_DIR / f"{stem}.png", dpi=600, bbox_inches="tight", facecolor="white")
    fig.savefig(
        FIG_DIR / f"{stem}.tiff",
        dpi=600,
        bbox_inches="tight",
        facecolor="white",
        pil_kwargs={"compression": "tiff_lzw"},
    )
    plt.close(fig)


def panel_label(ax: plt.Axes, label: str, x: float = -0.08, y: float = 1.04) -> None:
    ax.text(
        x,
        y,
        label,
        transform=ax.transAxes,
        ha="left",
        va="bottom",
        fontsize=10,
        fontweight="bold",
        color=COLORS["dark"],
    )


def clean_axes(ax: plt.Axes, grid_axis: str = "y") -> None:
    ax.grid(axis=grid_axis, color="#D9DEE1", linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["left"].set_color("#7F8A90")
    ax.spines["bottom"].set_color("#7F8A90")


def draw_box(
    ax: plt.Axes,
    x: float,
    y: float,
    width: float,
    height: float,
    title: str,
    lines: list[str],
    facecolor: str,
    title_size: float = 8.2,
    body_size: float = 7.1,
) -> None:
    box = FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.008,rounding_size=0.012",
        linewidth=1.0,
        edgecolor="#42545C",
        facecolor=facecolor,
    )
    ax.add_patch(box)
    ax.text(
        x + width / 2,
        y + height * 0.72,
        title,
        ha="center",
        va="center",
        fontsize=title_size,
        fontweight="bold",
        color="#17385E",
    )
    ax.text(
        x + width / 2,
        y + height * 0.36,
        "\n".join(lines),
        ha="center",
        va="center",
        fontsize=body_size,
        color="#20272B",
        linespacing=1.35,
    )


def arrow(ax: plt.Axes, start: tuple[float, float], end: tuple[float, float], color: str) -> None:
    ax.add_patch(
        FancyArrowPatch(
            start,
            end,
            arrowstyle="-|>",
            mutation_scale=10,
            linewidth=1.2,
            color=color,
            connectionstyle="arc3,rad=0",
        )
    )


def figure_2_1_conceptual_framework() -> None:
    fig, ax = plt.subplots(figsize=(7.2, 4.05))
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    ax.text(
        0.01,
        0.965,
        "Conceptual basis of archive-based metal-ratio and flow screening",
        ha="left",
        va="top",
        fontsize=11,
        fontweight="bold",
        color=COLORS["dark"],
    )
    ax.text(
        0.01,
        0.905,
        "Observed ratios reflect environmental processes and the way those processes are sampled.",
        ha="left",
        va="top",
        fontsize=8,
        color="#59646A",
    )

    top_y, top_h, top_w = 0.50, 0.29, 0.17
    xs = [0.01, 0.21, 0.41, 0.61, 0.81]
    titles = [
        "Mine sources",
        "Hydrological pathways",
        "In-stream processes",
        "Observed chemistry",
        "Screening output",
    ]
    bodies = [
        ["Ore mineralogy", "Waste and tailings", "Mine-water drainage"],
        ["Adit discharge", "Diffuse groundwater", "Runoff and erosion"],
        ["Dilution and mixing", "Sorption / precipitation", "Sediment mobilisation"],
        ["Pb, Zn and Cu", "Metal ratios", "pH and Ca context"],
        ["Repeated events", "Robust stations", "Field priorities"],
    ]
    fills = ["#D9E9E3", "#DCE7F3", "#F3E1D8", "#E8E0F0", "#F3E8B9"]
    for x, title, body, fill in zip(xs, titles, bodies, fills):
        draw_box(ax, x, top_y, top_w, top_h, title, body, fill)
    for left, right in zip(xs[:-1], xs[1:]):
        arrow(ax, (left + top_w + 0.006, top_y + top_h / 2), (right - 0.006, top_y + top_h / 2), COLORS["dark"])

    bottom_specs = [
        (0.19, "Hydrological context", ["Station-specific flow percentile", "Same-day, 3-day and 7-day windows"], "#DCE7F3"),
        (0.43, "Observation decisions", ["Sampling frequency and coverage", "Qualifiers and reporting limits"], "#E7E7E7"),
        (0.67, "Validation", ["Censoring sensitivity", "Manual gauge review"], "#F3E8B9"),
    ]
    for x, title, body, fill in bottom_specs:
        draw_box(ax, x, 0.10, 0.22, 0.23, title, body, fill, 8.0, 6.8)
    arrow(ax, (0.30, 0.33), (0.295, 0.49), COLORS["orange"])
    arrow(ax, (0.54, 0.33), (0.695, 0.49), COLORS["orange"])
    arrow(ax, (0.65, 0.215), (0.665, 0.215), COLORS["orange"])
    arrow(ax, (0.78, 0.33), (0.895, 0.49), COLORS["orange"])

    fig.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.02)
    save_figure(fig, "Figure_2_1_Conceptual_Framework")


def figure_3_1_study_area() -> None:
    wide = pd.read_csv(ROOT / "data/processed/mine_related_stations_wide.csv", low_memory=False)
    stations = (
        wide.dropna(subset=["easting", "northing"])
        .sort_values("sample_date")
        .groupby("station_id", as_index=False)
        .first()[["station_id", "station_name", "easting", "northing"]]
    )
    robust = pd.read_csv(ROOT / "results/7.6_site_profiles/7.6_Station_Flow_Map.csv")
    validation = pd.read_csv(ROOT / "data/processed/manual_flow_matching_validation_sites.csv")
    gauges = (
        validation[validation["station_id"].isin(robust["station_id"])]
        .dropna(subset=["manual_nrfa_easting", "manual_nrfa_northing"])
        .drop_duplicates("manual_nrfa_id")
    )

    systems = list(robust["system"].drop_duplicates())
    system_colors = {
        systems[0]: COLORS["red"],
        systems[1]: COLORS["green"],
        systems[2]: COLORS["purple"],
        systems[3]: COLORS["orange"],
    }

    fig = plt.figure(figsize=(7.2, 4.5))
    gs = fig.add_gridspec(1, 2, width_ratios=[1.05, 1.35], wspace=0.27)
    ax_all = fig.add_subplot(gs[0, 0])
    ax_zoom = fig.add_subplot(gs[0, 1])

    for ax in (ax_all, ax_zoom):
        ax.scatter(
            stations["easting"],
            stations["northing"],
            s=5,
            color=COLORS["gray"],
            alpha=0.35,
            edgecolors="none",
            zorder=1,
        )
        for system, group in robust.groupby("system"):
            ax.scatter(
                group["easting"],
                group["northing"],
                s=34,
                color=system_colors[system],
                edgecolor="white",
                linewidth=0.6,
                zorder=4,
                label=system,
            )
        ax.scatter(
            gauges["manual_nrfa_easting"],
            gauges["manual_nrfa_northing"],
            marker="^",
            s=45,
            color=COLORS["dark"],
            edgecolor="white",
            linewidth=0.5,
            zorder=5,
        )
        ax.set_aspect("equal", adjustable="box")
        clean_axes(ax, "both")
        ax.set_xlabel("British National Grid easting (m)")

    ax_all.set_ylabel("British National Grid northing (m)")
    ax_all.set_title("Wales-wide mine-related station coverage", loc="left", pad=10)
    ax_zoom.set_title("Robust stations and manually selected gauges", loc="left", pad=10)
    ax_zoom.set_xlim(238000, 286500)
    ax_zoom.set_ylim(235000, 296500)
    ax_zoom.tick_params(axis="y", labelleft=False)

    robust_with_gauge = robust.merge(
        validation[
            [
                "station_id",
                "manual_nrfa_easting",
                "manual_nrfa_northing",
            ]
        ],
        on="station_id",
        how="left",
    )
    for row in robust_with_gauge.itertuples():
        ax_zoom.plot(
            [row.easting, row.manual_nrfa_easting],
            [row.northing, row.manual_nrfa_northing],
            color="#8C989E",
            linewidth=0.7,
            alpha=0.65,
            zorder=2,
        )
    # Several stations are only tens of metres apart at the displayed scale.
    # Grouped callouts keep the map legible; Table 4.5 retains station-level detail.
    callouts = [
        (["S83017", "S83018", "S83019", "S83020", "S83021"], "S83017-S83021", (278300, 265900)),
        (["S35279", "S6320066"], "S35279 / S6320066", (277100, 293300)),
        (["S35767"], "S35767", (281200, 276300)),
        (["S35582"], "S35582", (272600, 276000)),
    ]
    for station_ids, label, text_position in callouts:
        group = robust[robust["station_id"].isin(station_ids)]
        x = group["easting"].mean()
        y = group["northing"].mean()
        system = group.iloc[0]["system"]
        ax_zoom.annotate(
            label,
            (x, y),
            xytext=text_position,
            fontsize=6.8,
            color=system_colors[system],
            ha="left",
            arrowprops={"arrowstyle": "-", "color": system_colors[system], "lw": 0.55},
            zorder=6,
        )

    handles = [
        Line2D([0], [0], marker="o", color="none", markerfacecolor=system_colors[s], markeredgecolor="white", markersize=6, label=s)
        for s in systems
    ]
    handles.append(
        Line2D([0], [0], marker="^", color="none", markerfacecolor=COLORS["dark"], markersize=6, label="Manually selected NRFA gauge")
    )
    fig.legend(handles=handles, loc="lower center", ncol=2, bbox_to_anchor=(0.53, -0.01))
    # Keep panel labels in a dedicated gutter to the left of each title.
    panel_label(ax_all, "a", x=-0.15, y=1.025)
    panel_label(ax_zoom, "b", x=-0.10, y=1.025)
    fig.subplots_adjust(left=0.09, right=0.99, top=0.885, bottom=0.20)
    save_figure(fig, "Figure_3_1_Study_Area_and_Flow_Gauges")

    robust_with_gauge.to_csv(SOURCE_DIR / "Figure_3_1_Source_Data.csv", index=False)


def figure_3_2_workflow() -> None:
    fig, ax = plt.subplots(figsize=(7.2, 3.55))
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    ax.text(
        0.01,
        0.96,
        "Stepwise narrowing from the NRW archive to robust site-level interpretation",
        ha="left",
        va="top",
        fontsize=10.5,
        fontweight="bold",
        color=COLORS["dark"],
    )

    steps = [
        ("Mine-related subset", ["595 stations", "20,178 station-date records"], "#DCE7F3"),
        ("Same-day Pb-Zn pairs", ["10,618 pairs", "438 stations"], "#D9E9E3"),
        ("Uncensored Pb-Zn", ["7,061 pairs", "288 stations"], "#F3E8B9"),
        ("Automatic flow linkage", ["5,627 same-day flows", "217 stations"], "#F3E1D8"),
        ("Robust interpretation", ["287 validated samples", "9 stations; 4 systems"], "#E8E0F0"),
    ]
    x_positions = [0.015, 0.21, 0.405, 0.60, 0.795]
    for x, (title, body, fill) in zip(x_positions, steps):
        draw_box(ax, x, 0.50, 0.175, 0.28, title, body, fill, 7.7, 7.0)
    for left, right in zip(x_positions[:-1], x_positions[1:]):
        arrow(ax, (left + 0.181, 0.64), (right - 0.006, 0.64), COLORS["dark"])

    checks = [
        (0.34, "Reporting-limit sensitivity", "As reported | half RL | uncensored"),
        (0.54, "Flow-window sensitivity", "Same day | 3-day | 7-day"),
        (0.74, "Hydrological validation", "30 stations manually reviewed"),
    ]
    for x, title, body in checks:
        draw_box(ax, x, 0.15, 0.18, 0.18, title, [body], "#F1F3F4", 7.2, 6.4)
    arrow(ax, (0.49, 0.50), (0.43, 0.34), COLORS["orange"])
    arrow(ax, (0.69, 0.50), (0.63, 0.34), COLORS["orange"])
    arrow(ax, (0.88, 0.50), (0.83, 0.34), COLORS["orange"])
    ax.text(
        0.015,
        0.08,
        "Interpretation rule: retain only patterns that remain credible under reasonable alternative processing choices.",
        fontsize=7.4,
        color="#59646A",
        ha="left",
    )
    fig.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.03)
    save_figure(fig, "Figure_3_2_Analytical_Workflow")


def figure_4_1_coverage_correlations() -> tuple[pd.DataFrame, pd.DataFrame]:
    coverage = pd.read_csv(ROOT / "data/processed/mine_related_parameter_coverage.csv")
    correlations = pd.read_csv(ROOT / "data/processed/mine_related_correlations.csv")
    order = ["lead", "zinc", "copper", "calcium", "ph", "hardness"]
    labels = ["Pb", "Zn", "Cu", "Ca", "pH", "Hardness"]
    coverage = coverage.set_index("parameter").reindex(order).reset_index()

    matrix_order = ["lead", "zinc", "copper", "calcium", "ph"]
    display = ["Pb", "Zn", "Cu", "Ca", "pH"]
    matrix = np.eye(len(matrix_order))
    sample_matrix = np.full_like(matrix, np.nan, dtype=float)
    for row in correlations.itertuples():
        i = matrix_order.index(row.parameter_1)
        j = matrix_order.index(row.parameter_2)
        matrix[i, j] = matrix[j, i] = row.spearman_r
        sample_matrix[i, j] = sample_matrix[j, i] = row.n_paired_samples

    fig = plt.figure(figsize=(7.2, 3.8))
    gs = fig.add_gridspec(1, 2, width_ratios=[0.85, 1.25], wspace=0.34)
    ax_cov = fig.add_subplot(gs[0, 0])
    ax_heat = fig.add_subplot(gs[0, 1])

    values = coverage["n_records"].fillna(0).to_numpy()
    colors = [COLORS["blue"]] * 5 + [COLORS["gray"]]
    bars = ax_cov.barh(labels[::-1], values[::-1], color=colors[::-1], edgecolor="white", zorder=2)
    for bar, value in zip(bars, values[::-1]):
        ax_cov.text(
            max(value, 0) + 350,
            bar.get_y() + bar.get_height() / 2,
            f"{int(value):,}",
            va="center",
            ha="left",
            fontsize=7.2,
            color=COLORS["dark"],
        )
    ax_cov.set_xlim(0, max(values) * 1.17)
    ax_cov.set_xlabel("Valid records")
    ax_cov.set_title("Archive coverage")
    clean_axes(ax_cov, "x")

    cmap = matplotlib.colors.LinearSegmentedColormap.from_list(
        "diverging",
        ["#3A6F9C", "#F7F7F7", "#B34743"],
    )
    im = ax_heat.imshow(matrix, vmin=-0.7, vmax=0.7, cmap=cmap, aspect="equal")
    ax_heat.set_xticks(range(5), display)
    ax_heat.set_yticks(range(5), display)
    ax_heat.set_title("Spearman rank correlations")
    for i in range(5):
        for j in range(5):
            value = matrix[i, j]
            color = "white" if abs(value) > 0.45 else COLORS["dark"]
            ax_heat.text(j, i, f"{value:.2f}", ha="center", va="center", fontsize=7.3, color=color)
    for spine in ax_heat.spines.values():
        spine.set_visible(False)
    cbar = fig.colorbar(im, ax=ax_heat, fraction=0.046, pad=0.04)
    cbar.set_label("Spearman rho", fontsize=8)
    cbar.ax.tick_params(labelsize=7)
    panel_label(ax_cov, "a")
    panel_label(ax_heat, "b")
    fig.subplots_adjust(left=0.08, right=0.96, top=0.87, bottom=0.15)
    save_figure(fig, "Figure_4_1_Coverage_and_Multi_Metal_Relationships")

    coverage.to_csv(SOURCE_DIR / "Figure_4_1a_Coverage.csv", index=False)
    correlations.to_csv(SOURCE_DIR / "Figure_4_1b_Correlations.csv", index=False)
    return coverage, correlations


def figure_4_2_sensitivity() -> tuple[pd.DataFrame, pd.DataFrame]:
    censor = pd.read_csv(ROOT / "results/7.5_sensitivity/7.5_Censoring_Metrics.csv")
    windows = pd.read_csv(ROOT / "results/7.5_sensitivity/7.5_Flow_Window_Metrics.csv")
    censor_labels = ["As reported", "Half reporting limit", "Uncensored only"]
    window_labels = ["Same day", "3-day mean", "7-day mean"]

    fig, axes = plt.subplots(1, 2, figsize=(7.2, 3.45), gridspec_kw={"wspace": 0.34})
    specs = [
        (
            axes[0],
            censor_labels,
            censor["n_candidates"].to_numpy(),
            censor["n_low_flow_high_ratio"].to_numpy(),
            censor["n_low_flow_high_pb_low_zn"].to_numpy(),
            [COLORS["blue"], COLORS["gold"], COLORS["green"]],
            "Reporting-limit treatment",
        ),
        (
            axes[1],
            window_labels,
            windows["n_candidates"].to_numpy(),
            windows["n_low_flow_high_ratio"].to_numpy(),
            windows["n_low_flow_high_pb_low_zn"].to_numpy(),
            [COLORS["orange"]] * 3,
            "Flow time window, uncensored pairs",
        ),
    ]
    for ax, labels_local, candidates, events, strict, colors_local, title in specs:
        x = np.arange(len(labels_local))
        bars = ax.bar(x, candidates, color=colors_local, edgecolor="white", width=0.68, zorder=2)
        for idx, (bar, candidate, event, strict_event) in enumerate(zip(bars, candidates, events, strict)):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                candidate + 0.35,
                f"{int(candidate)} stations",
                ha="center",
                va="bottom",
                fontsize=7.3,
                fontweight="bold",
            )
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                candidate * 0.45,
                f"{int(event)} primary\n{int(strict_event)} strict",
                ha="center",
                va="center",
                fontsize=7,
                color="white" if idx != 1 or ax is axes[1] else COLORS["dark"],
                fontweight="bold",
            )
        ax.set_xticks(x, labels_local, rotation=18, ha="right", rotation_mode="anchor")
        ax.set_ylabel("Candidate stations")
        ax.set_title(title)
        ax.set_ylim(0, max(candidates) + 3)
        clean_axes(ax, "y")
    panel_label(axes[0], "a")
    panel_label(axes[1], "b")
    fig.subplots_adjust(left=0.08, right=0.99, top=0.87, bottom=0.25)
    save_figure(fig, "Figure_4_2_Censoring_and_Flow_Window_Sensitivity")
    censor.to_csv(SOURCE_DIR / "Figure_4_2a_Censoring_Sensitivity.csv", index=False)
    windows.to_csv(SOURCE_DIR / "Figure_4_2b_Flow_Window_Sensitivity.csv", index=False)
    return censor, windows


def figure_4_3_matching_validation() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    distance = pd.read_csv(ROOT / "data/processed/manual_matching_agreement_by_distance.csv")
    samples = pd.read_csv(ROOT / "data/processed/pbzn_manual_flow_validation_samples.csv", low_memory=False)
    paired = samples[samples["both_methods_have_flow"].fillna(False)].copy()
    paired = paired.dropna(subset=["flow_percentile", "manual_flow_percentile"])
    metrics = pd.read_csv(ROOT / "data/processed/manual_matching_validation_metrics.csv")

    fig = plt.figure(figsize=(7.2, 4.15))
    gs = fig.add_gridspec(1, 3, width_ratios=[0.9, 1.25, 0.9], wspace=0.42)
    ax_distance = fig.add_subplot(gs[0, 0])
    ax_scatter = fig.add_subplot(gs[0, 1])
    ax_outcome = fig.add_subplot(gs[0, 2])

    distance_colors = [COLORS["green"], COLORS["gold"], COLORS["red"]]
    bars = ax_distance.bar(
        np.arange(len(distance)),
        distance["agreement_rate"] * 100,
        color=distance_colors,
        edgecolor="white",
        zorder=2,
    )
    for bar, row in zip(bars, distance.itertuples()):
        ax_distance.text(
            bar.get_x() + bar.get_width() / 2,
            row.agreement_rate * 100 + 3,
            f"{row.n_exact_agreement}/{row.n_sites}",
            ha="center",
            va="bottom",
            fontsize=7.3,
            fontweight="bold",
        )
    ax_distance.set_xticks(np.arange(len(distance)), ["<2 km", "2-<7 km", ">=7 km"])
    ax_distance.set_ylim(0, 112)
    ax_distance.set_ylabel("Exact gauge agreement (%)")
    ax_distance.set_title("Agreement by distance")
    clean_axes(ax_distance, "y")

    changed = paired["low_flow_status_changed"].fillna(False).astype(bool)
    ax_scatter.scatter(
        paired.loc[~changed, "flow_percentile"],
        paired.loc[~changed, "manual_flow_percentile"],
        s=9,
        color=COLORS["blue"],
        alpha=0.28,
        edgecolors="none",
        rasterized=True,
        label="Classification unchanged",
    )
    ax_scatter.scatter(
        paired.loc[changed, "flow_percentile"],
        paired.loc[changed, "manual_flow_percentile"],
        s=11,
        color=COLORS["red"],
        alpha=0.60,
        edgecolors="none",
        rasterized=True,
        label="Low-flow class changed",
    )
    ax_scatter.plot([0, 1], [0, 1], color=COLORS["dark"], linestyle="--", linewidth=0.9)
    ax_scatter.axvline(0.25, color="#7F8A90", linestyle=":", linewidth=0.8)
    ax_scatter.axhline(0.25, color="#7F8A90", linestyle=":", linewidth=0.8)
    ax_scatter.set_xlim(0, 1)
    ax_scatter.set_ylim(0, 1)
    ax_scatter.set_xlabel("Automatic flow percentile")
    ax_scatter.set_ylabel("Manual flow percentile")
    ax_scatter.set_title("Flow-percentile comparison")
    ax_scatter.text(
        0.04,
        0.94,
        "rho = 0.874\nn = 1,335\n8.99% changed class",
        transform=ax_scatter.transAxes,
        ha="left",
        va="top",
        fontsize=7.2,
        bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "#AAB2B6", "alpha": 0.92},
    )
    ax_scatter.legend(loc="lower right", fontsize=6.5)
    clean_axes(ax_scatter, "both")

    outcomes = pd.DataFrame(
        {
            "Outcome": ["Retained", "Disqualified", "Not evaluable"],
            "Sites": [17, 3, 1],
        }
    )
    outcome_colors = [COLORS["green"], COLORS["red"], COLORS["gray"]]
    outcome_bars = ax_outcome.barh(
        outcomes["Outcome"][::-1],
        outcomes["Sites"][::-1],
        color=outcome_colors[::-1],
        edgecolor="white",
        zorder=2,
    )
    for bar, value in zip(outcome_bars, outcomes["Sites"][::-1]):
        ax_outcome.text(value + 0.25, bar.get_y() + bar.get_height() / 2, str(value), va="center", fontsize=7.4, fontweight="bold")
    ax_outcome.set_xlim(0, 19)
    ax_outcome.set_xlabel("Initial candidate stations")
    ax_outcome.set_title("Candidate status")
    ax_outcome.text(
        0.02,
        0.04,
        "No new candidates\namong 9 controls",
        transform=ax_outcome.transAxes,
        fontsize=7,
        color="#59646A",
        ha="left",
        va="bottom",
    )
    clean_axes(ax_outcome, "x")

    panel_label(ax_distance, "a", -0.16)
    panel_label(ax_scatter, "b", -0.13)
    panel_label(ax_outcome, "c", -0.18)
    fig.subplots_adjust(left=0.08, right=0.99, top=0.86, bottom=0.18)
    save_figure(fig, "Figure_4_3_Manual_Flow_Matching_Validation")
    distance.to_csv(SOURCE_DIR / "Figure_4_3a_Agreement_By_Distance.csv", index=False)
    paired[
        [
            "station_id",
            "sample_date",
            "flow_percentile",
            "manual_flow_percentile",
            "low_flow_status_changed",
        ]
    ].to_csv(SOURCE_DIR / "Figure_4_3b_Flow_Percentile_Comparison.csv", index=False)
    outcomes.to_csv(SOURCE_DIR / "Figure_4_3c_Candidate_Status.csv", index=False)
    return distance, metrics, outcomes


def figure_4_4_site_heterogeneity() -> tuple[pd.DataFrame, pd.DataFrame]:
    sites = pd.read_csv(ROOT / "results/7.6_site_profiles/7.6_Site_Summary.csv")
    events = pd.read_csv(ROOT / "results/7.6_site_profiles/7.6_Event_Drivers.csv")
    sites = sites.sort_values("spearman_ratio_vs_flow_rho")

    fig = plt.figure(figsize=(7.2, 4.8))
    gs = fig.add_gridspec(1, 2, width_ratios=[1.15, 1.0], wspace=0.38)
    ax_rho = fig.add_subplot(gs[0, 0])
    ax_events = fig.add_subplot(gs[0, 1])

    y = np.arange(len(sites))
    significant = sites["spearman_ratio_vs_flow_fdr_q"] < 0.05
    point_colors = np.where(sites["spearman_ratio_vs_flow_rho"] < 0, COLORS["green"], COLORS["orange"])
    ax_rho.axvline(0, color="#727C81", linestyle="--", linewidth=0.9)
    for yi, row, color, sig in zip(y, sites.itertuples(), point_colors, significant):
        ax_rho.scatter(
            row.spearman_ratio_vs_flow_rho,
            yi,
            s=62 if sig else 38,
            facecolor=color,
            edgecolor="black" if sig else "white",
            linewidth=1.1 if sig else 0.5,
            zorder=3,
        )
    labels = [f"{row.station_id} (n={int(row.n_ratio_flow_correlation)})" for row in sites.itertuples()]
    ax_rho.set_yticks(y, labels)
    ax_rho.set_xlim(-0.75, 0.75)
    ax_rho.set_xlabel("Spearman rho: Pb/Zn versus flow percentile")
    ax_rho.set_title("Station-specific associations")
    ax_rho.text(
        0.02,
        0.98,
        "Pooled nine-station result:\nrho = 0.009; p = 0.883",
        transform=ax_rho.transAxes,
        ha="left",
        va="top",
        fontsize=7.2,
        bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "#AAB2B6", "alpha": 0.94},
    )
    clean_axes(ax_rho, "x")
    rho_legend = [
        Line2D([0], [0], marker="o", color="none", markerfacecolor=COLORS["green"], markeredgecolor="white", markersize=6, label="Negative rho"),
        Line2D([0], [0], marker="o", color="none", markerfacecolor=COLORS["orange"], markeredgecolor="white", markersize=6, label="Positive rho"),
        Line2D([0], [0], marker="o", color="none", markerfacecolor="white", markeredgecolor="black", markersize=7, label="FDR q < 0.05"),
    ]
    ax_rho.legend(handles=rho_legend, loc="lower right", fontsize=6.6)

    category_order = ["High Pb and low Zn", "High Pb only", "Low Zn only"]
    system_order = [
        "Esgair Mwyn / Nant y Garw",
        "Nant y Watcyn",
        "Esgair Hir / Eastern Ffraith",
        "Wemyss",
    ]
    event_pivot = (
        events[events["driver_category"].isin(category_order)]
        .pivot_table(index="system", columns="driver_category", values="n_events", aggfunc="sum", fill_value=0)
        .reindex(system_order)
        .reindex(columns=category_order, fill_value=0)
    )
    display_systems = ["Esgair Mwyn /\nNant y Garw", "Nant y Watcyn", "Esgair Hir /\nEastern Ffraith", "Wemyss"]
    left = np.zeros(len(event_pivot))
    event_colors = [COLORS["red"], COLORS["orange"], COLORS["blue"]]
    for category, color in zip(category_order, event_colors):
        values = event_pivot[category].to_numpy()
        bars = ax_events.barh(
            np.arange(len(event_pivot)),
            values,
            left=left,
            color=color,
            edgecolor="white",
            label=category,
            zorder=2,
        )
        for bar, value, offset in zip(bars, values, left):
            if value > 0:
                ax_events.text(
                    offset + value / 2,
                    bar.get_y() + bar.get_height() / 2,
                    str(int(value)),
                    ha="center",
                    va="center",
                    color="white",
                    fontsize=7.2,
                    fontweight="bold",
                )
        left += values
    ax_events.set_yticks(np.arange(len(display_systems)), display_systems)
    ax_events.invert_yaxis()
    ax_events.set_xlabel("Primary events (n = 49)")
    ax_events.set_title("Event composition by mine system")
    ax_events.legend(loc="lower right", fontsize=6.5)
    clean_axes(ax_events, "x")

    panel_label(ax_rho, "a", -0.13)
    panel_label(ax_events, "b", -0.14)
    fig.subplots_adjust(left=0.13, right=0.99, top=0.88, bottom=0.15)
    save_figure(fig, "Figure_4_4_Site_Heterogeneity_and_Event_Composition")
    sites.to_csv(SOURCE_DIR / "Figure_4_4a_Site_Correlations.csv", index=False)
    event_pivot.reset_index().to_csv(SOURCE_DIR / "Figure_4_4b_Event_Composition.csv", index=False)
    return sites, events


def figure_4_5_s83020_other_metals() -> tuple[pd.DataFrame, pd.DataFrame]:
    event_context = pd.read_csv(ROOT / "results/7.7_other_metals/7.7_PbZn_Event_Copper_Context.csv")
    s83020 = event_context[event_context["station_id"] == "S83020"].iloc[0]
    ratio_context = pd.read_csv(ROOT / "results/7.7_other_metals/7.7_Final_Nine_Site_Context.csv")
    ratio_s83020 = ratio_context[ratio_context["station_id"] == "S83020"].copy()
    ratio_s83020["pair"] = pd.Categorical(ratio_s83020["pair"], ["Pb/Zn", "Cu/Zn", "Pb/Cu"], ordered=True)
    ratio_s83020 = ratio_s83020.sort_values("pair")

    fold_labels = ["Pb", "Zn", "Cu", "Pb/Zn", "Cu/Zn", "Pb/Cu"]
    folds = np.array(
        [
            s83020.event_to_normal_lead_fold,
            s83020.event_to_normal_zinc_fold,
            s83020.event_to_normal_copper_fold,
            s83020.event_to_normal_pb_zn_ratio_fold,
            s83020.event_to_normal_cu_zn_ratio_fold,
            s83020.event_to_normal_pb_cu_ratio_fold,
        ]
    )
    fold_colors = [COLORS["red"], COLORS["blue"], COLORS["orange"], COLORS["purple"], COLORS["gold"], COLORS["gray"]]

    fig = plt.figure(figsize=(7.2, 3.75))
    gs = fig.add_gridspec(1, 2, width_ratios=[1.35, 0.95], wspace=0.38)
    ax_fold = fig.add_subplot(gs[0, 0])
    ax_rho = fig.add_subplot(gs[0, 1])

    x = np.arange(len(fold_labels))
    bars = ax_fold.bar(x, folds, color=fold_colors, edgecolor="white", zorder=2)
    ax_fold.axhline(1, color=COLORS["dark"], linestyle="--", linewidth=0.9)
    for bar, value in zip(bars, folds):
        ax_fold.text(
            bar.get_x() + bar.get_width() / 2,
            value + 0.06,
            f"{value:.2f}x",
            ha="center",
            va="bottom",
            fontsize=7.1,
            fontweight="bold",
        )
    ax_fold.set_xticks(x, fold_labels)
    ax_fold.set_ylim(0, 2.75)
    ax_fold.set_ylabel("Primary-event / normal-high-flow median")
    ax_fold.set_title("S83020 primary-event composition", loc="left", pad=12)
    clean_axes(ax_fold, "y")

    rho_values = ratio_s83020["spearman_ratio_vs_flow_rho"].to_numpy()
    q_values = ratio_s83020["spearman_ratio_vs_flow_fdr_q"].to_numpy()
    y = np.arange(len(ratio_s83020))
    ax_rho.axvline(0, color="#727C81", linestyle="--", linewidth=0.9)
    for yi, rho, q, color in zip(y, rho_values, q_values, [COLORS["purple"], COLORS["gold"], COLORS["gray"]]):
        ax_rho.scatter(
            rho,
            yi,
            s=66 if q < 0.05 else 42,
            facecolor=color,
            edgecolor="black" if q < 0.05 else "white",
            linewidth=1.1 if q < 0.05 else 0.5,
            zorder=3,
        )
        if rho > -0.25:
            text_x, text_ha = rho - 0.035, "right"
        else:
            text_x, text_ha = rho + 0.035, "left"
        ax_rho.text(text_x, yi, f"rho={rho:.3f}\nq={q:.3g}", ha=text_ha, va="center", fontsize=6.8)
    ax_rho.set_yticks(y, ratio_s83020["pair"].astype(str))
    ax_rho.invert_yaxis()
    ax_rho.set_xlim(-0.75, 0.15)
    ax_rho.set_xlabel("Spearman rho versus flow percentile")
    ax_rho.set_title("Ratio-flow associations at S83020", loc="left", pad=12)
    clean_axes(ax_rho, "x")
    panel_label(ax_fold, "a", x=-0.13, y=1.025)
    panel_label(ax_rho, "b", x=-0.18, y=1.025)
    fig.subplots_adjust(left=0.09, right=0.98, top=0.85, bottom=0.17)
    save_figure(fig, "Figure_4_5_S83020_Multi_Metal_Interpretation")

    pd.DataFrame({"measure": fold_labels, "event_to_normal_fold": folds}).to_csv(
        SOURCE_DIR / "Figure_4_5a_S83020_Event_Folds.csv", index=False
    )
    ratio_s83020.to_csv(SOURCE_DIR / "Figure_4_5b_S83020_Ratio_Flow_Correlations.csv", index=False)
    return event_context, ratio_context


def write_dataframe_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=title[:31])
    header_fill = PatternFill("solid", fgColor="3B6F8E")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    body_font = Font(name="Arial", size=9)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=str(column_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        values = [str(column_name)] + ["" if pd.isna(v) else str(v) for v in dataframe.iloc[:, column_index - 1].head(100)]
        width = min(max(max(len(v) for v in values) + 2, 11), 42)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.row_dimensions[1].height = 30


def build_tables(
    coverage: pd.DataFrame,
    correlations: pd.DataFrame,
    censor: pd.DataFrame,
    windows: pd.DataFrame,
    matching_metrics: pd.DataFrame,
    sites: pd.DataFrame,
    event_context: pd.DataFrame,
) -> None:
    wide = pd.read_csv(ROOT / "data/processed/mine_related_stations_wide.csv", low_memory=False)
    q = wide[["lead", "zinc", "copper", "calcium", "ph", "hardness"]].quantile([0.25, 0.75]).T
    units = {
        "lead": "micrograms/L",
        "zinc": "micrograms/L",
        "copper": "micrograms/L",
        "calcium": "mg/L",
        "ph": "pH units",
        "hardness": "-",
    }
    roles = {
        "lead": "Primary ratio numerator",
        "zinc": "Primary ratio denominator",
        "copper": "Extended ratio interpretation",
        "calcium": "Hydrogeochemical context",
        "ph": "Acidity context",
        "hardness": "Data gap",
    }
    table_4_1 = coverage[["parameter", "n_records", "n_stations", "median"]].copy()
    table_4_1["Q1"] = table_4_1["parameter"].map(q[0.25])
    table_4_1["Q3"] = table_4_1["parameter"].map(q[0.75])
    table_4_1["unit"] = table_4_1["parameter"].map(units)
    table_4_1["role_in_study"] = table_4_1["parameter"].map(roles)
    table_4_1["parameter"] = table_4_1["parameter"].map(
        {"lead": "Pb", "zinc": "Zn", "copper": "Cu", "calcium": "Ca", "ph": "pH", "hardness": "Hardness"}
    )
    table_4_1.columns = ["Indicator", "Valid records", "Stations", "Median", "Q1", "Q3", "Unit", "Role in study"]

    table_4_2 = correlations[["parameter_1", "parameter_2", "n_paired_samples", "spearman_r", "spearman_p"]].copy()
    display_map = {"lead": "Pb", "zinc": "Zn", "copper": "Cu", "calcium": "Ca", "ph": "pH"}
    table_4_2["Variable pair"] = table_4_2["parameter_1"].map(display_map) + "-" + table_4_2["parameter_2"].map(display_map)
    table_4_2 = table_4_2[["Variable pair", "n_paired_samples", "spearman_r", "spearman_p"]]
    table_4_2.columns = ["Variable pair", "Paired samples", "Spearman rho", "Two-sided p value"]

    censor_table = censor[
        [
            "scenario",
            "n_pbzn_samples",
            "n_same_day_flow",
            "spearman_ratio_flow",
            "n_low_flow_high_ratio",
            "n_low_flow_high_pb_low_zn",
            "n_candidates",
        ]
    ].copy()
    censor_table.insert(0, "Sensitivity family", "Censoring treatment")
    censor_table.columns = [
        "Sensitivity family",
        "Scenario",
        "Pb-Zn pairs",
        "Flow-matched samples",
        "Ratio-flow rho",
        "Primary events",
        "Strict events",
        "Candidate stations",
    ]
    window_table = windows[
        [
            "method",
            "n_with_flow",
            "spearman_ratio_flow",
            "n_low_flow_high_ratio",
            "n_low_flow_high_pb_low_zn",
            "n_candidates",
            "low_flow_class_change_rate",
        ]
    ].copy()
    window_table.insert(0, "Sensitivity family", "Flow window")
    window_table.insert(2, "Pb-Zn pairs", np.nan)
    window_table.columns = [
        "Sensitivity family",
        "Scenario",
        "Pb-Zn pairs",
        "Flow-matched samples",
        "Ratio-flow rho",
        "Primary events",
        "Strict events",
        "Candidate stations",
        "Low-flow class change rate",
    ]
    censor_table["Low-flow class change rate"] = np.nan
    table_4_3 = pd.concat([censor_table, window_table], ignore_index=True)

    table_4_4 = matching_metrics.copy()
    table_4_4.columns = ["Validation metric", "Value", "Unit"]

    table_4_5 = sites[
        [
            "system",
            "station_id",
            "station_name",
            "n_with_manual_flow",
            "n_low_flow_high_ratio_events",
            "n_strict_high_pb_low_zn_events",
            "spearman_ratio_vs_flow_rho",
            "spearman_ratio_vs_flow_p",
            "spearman_ratio_vs_flow_fdr_q",
            "flow_hypothesis_interpretation",
        ]
    ].copy()
    table_4_5.columns = [
        "Mine system",
        "Station ID",
        "Station name",
        "n with manual flow",
        "Primary events",
        "Strict events",
        "Spearman rho",
        "p value",
        "FDR q value",
        "Interpretation",
    ]

    s83020 = event_context[event_context["station_id"] == "S83020"].iloc[0]
    table_4_6 = pd.DataFrame(
        {
            "Measure": ["Pb", "Zn", "Cu", "Pb/Zn", "Cu/Zn", "Pb/Cu"],
            "Primary-event median": [
                s83020.event_median_lead,
                s83020.event_median_zinc,
                s83020.event_median_copper,
                s83020.event_median_pb_zn_ratio,
                s83020.event_median_cu_zn_ratio,
                s83020.event_median_pb_cu_ratio,
            ],
            "Normal/high-flow median": [
                s83020.normal_median_lead,
                s83020.normal_median_zinc,
                s83020.normal_median_copper,
                s83020.normal_median_pb_zn_ratio,
                s83020.normal_median_cu_zn_ratio,
                s83020.normal_median_pb_cu_ratio,
            ],
            "Fold change": [
                s83020.event_to_normal_lead_fold,
                s83020.event_to_normal_zinc_fold,
                s83020.event_to_normal_copper_fold,
                s83020.event_to_normal_pb_zn_ratio_fold,
                s83020.event_to_normal_cu_zn_ratio_fold,
                s83020.event_to_normal_pb_cu_ratio_fold,
            ],
            "Unit": ["micrograms/L", "micrograms/L", "micrograms/L", "dimensionless", "dimensionless", "dimensionless"],
        }
    )

    optional_table_paths = [
        (
            "Table 2.1 Literature",
            "2_1_Literature_Synthesis",
            ROOT / "results/thesis/chapters_1_3/Table_2_1_Literature_Synthesis.csv",
        ),
        (
            "Table 3.1 Data sources",
            "3_1_Data_Sources",
            ROOT / "results/thesis/chapters_1_3/Table_3_1_Data_Sources.csv",
        ),
        (
            "Table 3.2 QC rules",
            "3_2_Quality_Control_Rules",
            ROOT / "results/thesis/chapters_1_3/Table_3_2_Quality_Control_Rules.csv",
        ),
    ]
    optional_tables = [
        (title, stem, pd.read_csv(path))
        for title, stem, path in optional_table_paths
        if path.exists()
    ]
    table_3_3 = pd.DataFrame(
        [
            ["Low flow", "Flow percentile <= 25", "Relative hydrological classification"],
            ["High Pb", "Pb >= 510 micrograms/L", "75th percentile of uncensored Pb"],
            ["Low Zn", "Zn <= 452 micrograms/L", "25th percentile of uncensored Zn"],
            ["High Pb/Zn", "Pb/Zn >= 0.5163", "90th percentile of uncensored Pb/Zn"],
            ["Primary event", "Low flow and high Pb/Zn", "Co-occurrence screen"],
            ["Strict event", "Primary event, high Pb and low Zn", "Separates ratio from absolute concentrations"],
            ["Candidate station", ">=2 strict events; or >=3 primary events; or >=2 primary events comprising >=30% of matched samples", "Regional screening rule"],
            ["Robust station", "Retained after uncensored, flow-window and manual-matching checks", "Final interpretation set"],
        ],
        columns=["Definition", "Operational rule", "Purpose"],
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook_tables = [(title, dataframe) for title, _, dataframe in optional_tables] + [
        ("Table 3.3 Definitions", table_3_3),
        ("Table 4.1 Coverage", table_4_1),
        ("Table 4.2 Correlations", table_4_2),
        ("Table 4.3 Sensitivity", table_4_3),
        ("Table 4.4 Validation", table_4_4),
        ("Table 4.5 Robust stations", table_4_5),
        ("Table 4.6 S83020", table_4_6),
    ]
    for title, dataframe in workbook_tables:
        write_dataframe_sheet(workbook, title, dataframe)
    workbook.save(TABLE_DIR / "Final_Dissertation_Tables.xlsx")

    csv_tables = [(stem, dataframe) for _, stem, dataframe in optional_tables] + [
        ("3_3_Operational_Definitions", table_3_3),
        ("4_1_Data_Coverage", table_4_1),
        ("4_2_Multi_Metal_Correlations", table_4_2),
        ("4_3_Sensitivity", table_4_3),
        ("4_4_Manual_Matching_Validation", table_4_4),
        ("4_5_Robust_Stations", table_4_5),
        ("4_6_S83020_Multi_Metal_Context", table_4_6),
    ]
    for number, dataframe in csv_tables:
        dataframe.to_csv(TABLE_DIR / f"Table_{number}.csv", index=False)


def copy_appendix_figures() -> None:
    appendix_dir = FIG_DIR / "appendix_options"
    appendix_dir.mkdir(exist_ok=True)
    files = {
        ROOT / "results/7.6_site_profiles/7.6_Time_Series.png": "Figure_A_1_Robust_Station_Time_Series.png",
        ROOT / "results/7.6_site_profiles/7.6_pH_Ca_Context.png": "Figure_A_2_pH_and_Ca_Context.png",
        ROOT / "results/7.7_other_metals/7.7_Ratio_vs_Flow.png": "Figure_A_3_Extended_Ratios_vs_Flow.png",
        ROOT / "results/7.2_mine_related/7.2_Scatterplots.png": "Figure_A_4_Multi_Metal_Scatterplots.png",
    }
    for source, destination in files.items():
        if source.exists():
            shutil.copy2(source, appendix_dir / destination)


def write_manifest() -> None:
    manifest = {
        "core_conclusion": "The regional workflow identifies robust site-specific signals, but it does not support a universal low-flow increase in Pb/Zn.",
        "backend": "Python (matplotlib)",
        "main_figures": [
            "Figure_2_1_Conceptual_Framework",
            "Figure_3_1_Study_Area_and_Flow_Gauges",
            "Figure_3_2_Analytical_Workflow",
            "Figure_4_1_Coverage_and_Multi_Metal_Relationships",
            "Figure_4_2_Censoring_and_Flow_Window_Sensitivity",
            "Figure_4_3_Manual_Flow_Matching_Validation",
            "Figure_4_4_Site_Heterogeneity_and_Event_Composition",
            "Figure_4_5_S83020_Multi_Metal_Interpretation",
        ],
        "data_integrity": {
            "coverage": "All six target indicators; hardness retained as a zero-coverage category.",
            "matching_scatter": "All 1,335 samples with flow available under automatic and manual matching.",
            "site_heterogeneity": "All nine robust stations and all 49 final primary events.",
            "S83020_event_comparison": "Three primary-event samples with Cu and 39 normal/high-flow samples with Cu.",
        },
        "exports": ["SVG with editable text", "PDF", "400 dpi PNG"],
    }
    (OUT / "figure_manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def main() -> None:
    figure_2_1_conceptual_framework()
    figure_3_1_study_area()
    figure_3_2_workflow()
    coverage, correlations = figure_4_1_coverage_correlations()
    censor, windows = figure_4_2_sensitivity()
    _, matching_metrics, _ = figure_4_3_matching_validation()
    sites, _ = figure_4_4_site_heterogeneity()
    event_context, _ = figure_4_5_s83020_other_metals()
    build_tables(coverage, correlations, censor, windows, matching_metrics, sites, event_context)
    copy_appendix_figures()
    write_manifest()
    print(f"Created figure and table pack at: {OUT}")


if __name__ == "__main__":
    main()
