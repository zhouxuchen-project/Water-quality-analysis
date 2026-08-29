from __future__ import annotations

import json
import math
import shutil
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PROCESSED = ROOT / "data" / "processed"
RAW_FLOW = ROOT / "data" / "raw" / "river_flow"
RESULTS = ROOT / "results" / "7.4_manual_matching_validation"
PROJECT_RESULTS = ROOT / "results" / "project_numbered"

NRFA_BASE = "https://nrfaapps.ceh.ac.uk/nrfa/ws"
LOW_FLOW_PERCENTILE = 0.25
VERY_LOW_FLOW_PERCENTILE = 0.10

CANDIDATE_IDS = [
    "S83019",
    "S35767",
    "S83018",
    "S30418",
    "S83017",
    "S17302",
    "S44028",
    "S17303",
    "S35279",
    "S6320066",
    "S35634",
    "S83020",
    "S88433",
    "S83021",
    "S35633",
    "S88224",
    "S35582",
    "S71107",
    "S32484",
    "S26480",
    "S26477",
]

CONTROL_IDS = [
    "S20247",  # short distance, River Mawddach
    "S71656",  # short distance, River Neath
    "S35628",  # short distance, River Ystwyth
    "S717",  # medium distance, River Alyn
    "S71081",  # medium distance, River Afan
    "S72934",  # medium distance, River Loughor
    "S81011",  # long distance, Rheidol/Ystwyth test
    "S26278",  # long distance, Afon Goch/Cefni test
    "S22826",  # long distance, Afon Barlwydd/Glaslyn test
]

SOURCES = {
    "nrfa": {
        "title": "NRFA API and station metadata",
        "url": "https://nrfaapps.ceh.ac.uk/nrfa/nrfa-api.html",
        "use": "Official gauging-station river, coordinates, period of record and daily mean flow.",
    },
    "esgair_mwyn": {
        "title": "NRW: Esgair Mwyn spoil-heap erosion prevention works",
        "url": "https://ymgynghori.cyfoethnaturiol.cymru/metal-mines/no-environmental-statement-esgair-mwyn/",
        "use": "Identifies Esgair Mwyn drainage to Nant y Garw; NRW case-study evidence links it through Afon Meurig to the Teifi system.",
    },
    "trelewis": {
        "title": "GOV.UK: Taff Merthyr mine-water treatment scheme",
        "url": "https://www.gov.uk/government/case-studies/taff-merthyr-mine-water-treatment-scheme",
        "use": "Identifies the Trelewis/Taff Merthyr discharge as entering the Taff Bargoed.",
    },
    "blaenavon": {
        "title": "Welsh Government: Mine-water opportunity mapping, Torfaen",
        "url": "https://www.gov.wales/sites/default/files/publications/2024-08/mine-water-heat-opportunities-torfaen.pdf",
        "use": "Identifies the Blaenavon treatment outflow as entering River Arch/Afon Lwyd.",
    },
    "esgair_hir": {
        "title": "Welsh Potosi: Esgair Hir and Esgair Ffriath surface-water management",
        "url": "https://www.gwerthwchigymru.llyw.cymru/search/show/search_view.aspx?ID=JUN623333&catID=",
        "use": "Identifies Nant Du and Afon Lluestgota as pathways to Nant-y-Moch Reservoir and the River Rheidol.",
    },
    "lindsay": {
        "title": "Welsh Government: Mine-water opportunity mapping, Carmarthenshire",
        "url": "https://www.gov.wales/sites/default/files/publications/2024-08/mine-water-heat-opportunities-camarthenshire.pdf",
        "use": "Identifies Lindsay treatment discharge to Fferrws Brook in the Loughor system.",
    },
    "nant_y_bai": {
        "title": "RSC: Nantymwyn hydrological pathways study",
        "url": "https://pubs.rsc.org/en/content/articlehtml/2024/ew/d3ew00194f",
        "use": "Identifies Nant y Bai as flowing through the mine and into the River Tywi.",
    },
    "corrwg": {
        "title": "NRW River Basin Management Plan annex",
        "url": "https://naturalresourceswales.gov.uk/media/675075/annex-d.pdf",
        "use": "Identifies Corrwg Fechan as part of the Corrwg/Afan river system.",
    },
    "ynysarwed": {
        "title": "Ynysarwed mine-water treatment study",
        "url": "https://orca.cardiff.ac.uk/id/eprint/38400/",
        "use": "Identifies the Ynysarwed discharge as affecting the Neath Canal and River Neath.",
    },
    "whitworth": {
        "title": "IMWA: Pelenna mine-water treatment sites",
        "url": "https://www.imwa.info/docs/imwa_2018/IMWA2018_Bannister_84.pdf",
        "use": "Identifies Whitworth/Pelenna drainage as a tributary of the River Afan.",
    },
    "parys": {
        "title": "Peer-reviewed study of Parys Mountain Afon Goch",
        "url": "https://www.frontiersin.org/journals/microbiology/articles/10.3389/fmicb.2018.01446/full",
        "use": "Identifies Dyffryn Adda Adit as entering the northern Afon Goch, which flows to the coast rather than the Cefni.",
    },
    "oakeley": {
        "title": "Wales Slate: Ffestiniog slate landscape",
        "url": "https://www.llechi.cymru/slateareas/ffestiniog",
        "use": "Supports the Ffestiniog/Oakeley setting; local station naming identifies Afon Barlwydd, for which no NRFA mean-flow gauge was found.",
    },
}


def mapping(
    manual_id: int | None,
    best_id: int | None,
    decision: str,
    confidence: str,
    evidence: str,
    source_key: str,
) -> dict[str, Any]:
    return {
        "manual_nrfa_id": manual_id,
        "hydrologic_best_nrfa_id": best_id,
        "manual_decision": decision,
        "manual_confidence": confidence,
        "hydrological_basis": evidence,
        "source_key": source_key,
    }


MANUAL_MATCHES: dict[str, dict[str, Any]] = {}

for station_id in ["S83017", "S83018", "S83019", "S83020", "S83021"]:
    MANUAL_MATCHES[station_id] = mapping(
        62002,
        62002,
        "Replace automatic match",
        "Medium",
        "Esgair Mwyn drains via Nant y Garw and Afon Meurig to the Teifi. Teifi at Llanfair is the available downstream NRFA mean-flow station, but it is distant and integrates a much larger catchment.",
        "esgair_mwyn",
    )

for station_id in ["S35767", "S35634", "S35633", "S32484", "S35628"]:
    MANUAL_MATCHES[station_id] = mapping(
        63004,
        63004,
        "Retain automatic match",
        "High",
        "The station is on the upper Ystwyth or an immediate tributary near Cwm Ystwyth; the Cwm Ystwyth gauge is on the same local receiving river.",
        "nrfa",
    )

MANUAL_MATCHES["S30418"] = mapping(
    59002,
    59002,
    "Retain automatic match",
    "Medium",
    "The Lindsay treatment system discharges to Fferrws Brook in the Loughor system. Tir-y-dail is the closest available main-river flow proxy, although it is not a tributary-specific gauge.",
    "lindsay",
)

for station_id in ["S17302", "S17303"]:
    MANUAL_MATCHES[station_id] = mapping(
        57007,
        57007,
        "Replace automatic match",
        "High",
        "The Trelewis/Taff Merthyr discharge enters the Taff Bargoed. Taff at Fiddlers Elbow is downstream in the connected Taff system; Cynon at Abercynon is on a different tributary.",
        "trelewis",
    )

MANUAL_MATCHES["S44028"] = mapping(
    56005,
    56005,
    "Hydrological station has no sample-period overlap",
    "High",
    "Blaenavon outflow enters River Arch/Afon Lwyd. Lwyd at Ponthir is hydrologically connected but closed in 1998, before the 2002-2005 water-quality samples; no overlapping exact NRFA gauge was available.",
    "blaenavon",
)

for station_id in ["S35279", "S6320066", "S81011"]:
    MANUAL_MATCHES[station_id] = mapping(
        63002,
        63002,
        "Replace automatic match",
        "High",
        "Esgair Hir/Esgair Ffriath and Cwm Rheidol drainage belongs to the Rheidol system. Rheidol at Llanbadarn Fawr is downstream; the automatically selected Dyfi or Ystwyth station is in another river system.",
        "esgair_hir",
    )

for station_id in ["S88433", "S88224"]:
    MANUAL_MATCHES[station_id] = mapping(
        60007,
        60007,
        "Replace automatic match",
        "High",
        "Nant y Bai flows into the River Tywi. Tywi at Dolau Hirion is downstream of the confluence; Ystradffin is closer but upstream of the mine tributary input.",
        "nant_y_bai",
    )

MANUAL_MATCHES["S35582"] = mapping(
    63001,
    63001,
    "Replace automatic match",
    "Medium",
    "Wemyss mine drainage affects the Ystwyth downstream of Cwm Ystwyth. Pont Llolwyn is the downstream Ystwyth gauge; the closer Cwm Ystwyth gauge is upstream of this mine input.",
    "nrfa",
)

MANUAL_MATCHES["S71107"] = mapping(
    58004,
    58004,
    "Replace automatic match",
    "High",
    "Afon Corrwg Fechan joins the Corrwg and then the River Afan. Afan at Cwmafan is downstream; Rhondda Fawr is across the watershed.",
    "corrwg",
)

for station_id in ["S26477", "S26480"]:
    MANUAL_MATCHES[station_id] = mapping(
        64001,
        64004,
        "Replace with downstream station having sample-period data",
        "Medium",
        "The site naming identifies the Afon Twymyn. Twymyn at Cemmaes Road is the best direct gauge but closed in 2001; Dyfi at Dyfi Bridge is the available downstream gauge for the 2007 samples.",
        "nrfa",
    )

MANUAL_MATCHES["S20247"] = mapping(
    64010,
    64010,
    "Retain automatic match",
    "High",
    "The water-quality station is explicitly on the River Mawddach and the nearby NRFA station is on the same river.",
    "nrfa",
)

MANUAL_MATCHES["S71656"] = mapping(
    58002,
    58002,
    "Retain automatic match",
    "High",
    "The Ynysarwed discharge affects the Neath Canal and River Neath; Neath at Resolven is the connected nearby gauge.",
    "ynysarwed",
)

MANUAL_MATCHES["S717"] = mapping(
    67009,
    67009,
    "Retain automatic match",
    "High",
    "The station is explicitly on the River Alyn and the matched NRFA station is on the Alyn.",
    "nrfa",
)

MANUAL_MATCHES["S71081"] = mapping(
    58004,
    58004,
    "Retain automatic match",
    "High",
    "Whitworth/Pelenna drainage enters the River Afan; Afan at Cwmafan is the connected downstream gauge.",
    "whitworth",
)

MANUAL_MATCHES["S72934"] = mapping(
    59002,
    59002,
    "Retain automatic match",
    "Medium",
    "Fferrws Brook lies in the Loughor system. Tir-y-dail is the closest available main-river proxy, with the same tributary-scale limitation as the Lindsay site.",
    "lindsay",
)

MANUAL_MATCHES["S26278"] = mapping(
    None,
    None,
    "No suitable NRFA mean-flow gauge",
    "High",
    "Dyffryn Adda Adit enters the northern Afon Goch, which flows directly to the coast. Cefni at Bodffordd is a different catchment, and no Afon Goch NRFA mean-flow gauge was found.",
    "parys",
)

MANUAL_MATCHES["S22826"] = mapping(
    None,
    None,
    "No suitable NRFA mean-flow gauge",
    "Medium",
    "The station is on Afon Barlwydd in the Ffestiniog/Dwyryd system. Glaslyn at Beddgelert is across the watershed, and no Barlwydd or Dwyryd NRFA mean-flow gauge was found.",
    "oakeley",
)


def ensure_directories() -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    PROJECT_RESULTS.mkdir(parents=True, exist_ok=True)


def distance_band(distance_km: float) -> str:
    if distance_km < 2:
        return "Short (<2 km)"
    if distance_km < 7:
        return "Medium (2-<7 km)"
    return "Long (>=7 km)"


def load_validation_sites() -> tuple[pd.DataFrame, pd.DataFrame]:
    matches = pd.read_csv(PROCESSED / "pbzn_nearest_flow_station_matches.csv")
    station_summary = pd.read_csv(PROCESSED / "pbzn_flow_station_summary.csv")
    flow_metadata = pd.read_csv(RAW_FLOW / "nrfa_station_metadata_all.csv")

    selected_ids = CANDIDATE_IDS + CONTROL_IDS
    selected = matches[matches["station_id"].isin(selected_ids)].copy()
    assert len(selected) == 30, f"Expected 30 validation sites, found {len(selected)}"
    assert set(selected_ids) == set(MANUAL_MATCHES), "Manual mapping table is incomplete"

    selected["selection_group"] = np.where(
        selected["station_id"].isin(CANDIDATE_IDS),
        "Systematic candidate",
        "Stratified control",
    )
    selected["automatic_distance_band"] = selected["flow_match_distance_km"].map(distance_band)
    selected = selected.merge(
        station_summary[
            [
                "station_id",
                "n_pbzn_samples",
                "n_with_flow",
                "n_low_flow_high_pbzn_ratio",
                "n_low_flow_high_pb_low_zinc",
                "median_pb_zn_ratio",
                "systematic_candidate",
            ]
        ],
        on="station_id",
        how="left",
    )

    manual_rows = [{"station_id": key, **value} for key, value in MANUAL_MATCHES.items()]
    selected = selected.merge(pd.DataFrame(manual_rows), on="station_id", how="left")

    metadata_cols = [
        "nrfa_station_id",
        "nrfa_station_name",
        "river",
        "easting",
        "northing",
        "gdf_start_date",
        "gdf_end_date",
        "gdf_percent_complete",
    ]
    manual_metadata = flow_metadata[metadata_cols].rename(
        columns={
            "nrfa_station_id": "manual_nrfa_id",
            "nrfa_station_name": "manual_nrfa_name",
            "river": "manual_nrfa_river",
            "easting": "manual_nrfa_easting",
            "northing": "manual_nrfa_northing",
            "gdf_start_date": "manual_gdf_start_date",
            "gdf_end_date": "manual_gdf_end_date",
            "gdf_percent_complete": "manual_gdf_percent_complete",
        }
    )
    selected = selected.merge(manual_metadata, on="manual_nrfa_id", how="left")

    best_metadata = flow_metadata[metadata_cols].rename(
        columns={
            "nrfa_station_id": "hydrologic_best_nrfa_id",
            "nrfa_station_name": "hydrologic_best_nrfa_name",
            "river": "hydrologic_best_river",
            "gdf_start_date": "hydrologic_best_start_date",
            "gdf_end_date": "hydrologic_best_end_date",
        }
    )[
        [
            "hydrologic_best_nrfa_id",
            "hydrologic_best_nrfa_name",
            "hydrologic_best_river",
            "hydrologic_best_start_date",
            "hydrologic_best_end_date",
        ]
    ]
    selected = selected.merge(best_metadata, on="hydrologic_best_nrfa_id", how="left")

    dx = selected["easting"] - selected["manual_nrfa_easting"]
    dy = selected["northing"] - selected["manual_nrfa_northing"]
    selected["manual_match_distance_km"] = np.sqrt(dx**2 + dy**2) / 1000
    selected["exact_station_agreement"] = (
        selected["manual_nrfa_id"].notna()
        & (selected["nrfa_station_id"].astype("Int64") == selected["manual_nrfa_id"].astype("Int64"))
    )
    selected["automatic_match_assessment"] = np.select(
        [
            selected["exact_station_agreement"],
            selected["manual_nrfa_id"].isna(),
        ],
        ["Acceptable", "Not acceptable; no suitable NRFA gauge"],
        default="Not acceptable; replace",
    )
    selected["evidence_title"] = selected["source_key"].map(lambda key: SOURCES[key]["title"])
    selected["evidence_url"] = selected["source_key"].map(lambda key: SOURCES[key]["url"])
    selected["nrfa_station_url"] = selected["manual_nrfa_id"].map(
        lambda value: "" if pd.isna(value) else f"https://nrfa.ceh.ac.uk/data/station/info/{int(value)}"
    )

    selected["selection_order"] = selected["station_id"].map(
        {station_id: idx for idx, station_id in enumerate(selected_ids)}
    )
    selected = selected.sort_values("selection_order").reset_index(drop=True)

    selected.to_csv(PROCESSED / "manual_flow_matching_validation_sites.csv", index=False)
    return selected, flow_metadata


def parse_nrfa_stream(payload: dict[str, Any]) -> pd.DataFrame:
    stream = payload.get("data-stream", [])
    station = payload.get("station", {})
    station_id = station.get("id")
    station_name = station.get("name")
    rows: list[dict[str, Any]] = []
    for idx in range(0, len(stream), 2):
        date_value = stream[idx]
        raw_value = stream[idx + 1]
        flag = ""
        if isinstance(raw_value, list):
            value = raw_value[0]
            flag = raw_value[1] if len(raw_value) > 1 else ""
        else:
            value = raw_value
        rows.append(
            {
                "nrfa_station_id": int(station_id),
                "nrfa_station_name": station_name,
                "flow_date": pd.Timestamp(date_value).floor("D"),
                "flow_m3s": pd.to_numeric(value, errors="coerce"),
                "flow_flag": flag,
            }
        )
    return pd.DataFrame(rows).dropna(subset=["flow_m3s"])


def load_manual_flows(station_ids: list[int]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for station_id in sorted(set(station_ids)):
        raw_path = RAW_FLOW / f"nrfa_gdf_{station_id}.json"
        if raw_path.exists():
            payload = json.loads(raw_path.read_text(encoding="utf-8"))
        else:
            url = (
                f"{NRFA_BASE}/time-series"
                f"?format=json-object&data-type=gdf&station={station_id}&flags=true&dates=true"
            )
            response = requests.get(url, timeout=120)
            response.raise_for_status()
            payload = response.json()
            raw_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        frame = parse_nrfa_stream(payload)
        if not frame.empty:
            frames.append(frame)
    flows = pd.concat(frames, ignore_index=True)
    flows.to_csv(PROCESSED / "manual_matching_daily_flow.csv", index=False)
    return flows


def percentile_of_score(sorted_values: np.ndarray, values: pd.Series) -> np.ndarray:
    positions = np.searchsorted(sorted_values, values.to_numpy(dtype=float), side="right")
    return positions / len(sorted_values)


def build_sample_comparison(
    selected: pd.DataFrame, flows: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    samples = pd.read_csv(PROCESSED / "pbzn_flow_matched_samples.csv", parse_dates=["sample_day", "flow_date"])
    samples = samples[samples["station_id"].isin(selected["station_id"])].copy()
    sample_cols = [
        "station_id",
        "selection_group",
        "manual_nrfa_id",
        "manual_nrfa_name",
        "manual_nrfa_river",
        "manual_match_distance_km",
        "manual_decision",
        "manual_confidence",
        "exact_station_agreement",
    ]
    samples = samples.merge(selected[sample_cols], on="station_id", how="left")
    samples["manual_nrfa_id"] = samples["manual_nrfa_id"].astype("Int64")

    manual_flow = flows.rename(
        columns={
            "nrfa_station_id": "manual_nrfa_id",
            "nrfa_station_name": "manual_flow_station_name",
            "flow_date": "manual_flow_date",
            "flow_m3s": "manual_flow_m3s",
            "flow_flag": "manual_flow_flag",
        }
    )
    samples = samples.merge(
        manual_flow,
        left_on=["manual_nrfa_id", "sample_day"],
        right_on=["manual_nrfa_id", "manual_flow_date"],
        how="left",
    )

    samples["manual_flow_percentile"] = np.nan
    samples["manual_flow_p10_m3s"] = np.nan
    samples["manual_flow_p25_m3s"] = np.nan
    samples["manual_flow_p50_m3s"] = np.nan
    for station_id, indices in samples.groupby("manual_nrfa_id", dropna=True).groups.items():
        full_record = flows.loc[
            flows["nrfa_station_id"] == int(station_id), "flow_m3s"
        ].dropna()
        full_record = full_record[full_record >= 0].sort_values().to_numpy()
        valid = samples.loc[indices, "manual_flow_m3s"].notna()
        valid_indices = samples.loc[indices].index[valid]
        if len(full_record) and len(valid_indices):
            samples.loc[valid_indices, "manual_flow_percentile"] = percentile_of_score(
                full_record,
                samples.loc[valid_indices, "manual_flow_m3s"],
            )
            samples.loc[indices, "manual_flow_p10_m3s"] = np.quantile(full_record, 0.10)
            samples.loc[indices, "manual_flow_p25_m3s"] = np.quantile(full_record, 0.25)
            samples.loc[indices, "manual_flow_p50_m3s"] = np.quantile(full_record, 0.50)

    samples["manual_has_flow_match"] = samples["manual_flow_m3s"].notna()
    samples["manual_low_flow"] = samples["manual_flow_percentile"] <= LOW_FLOW_PERCENTILE
    samples["manual_very_low_flow"] = samples["manual_flow_percentile"] <= VERY_LOW_FLOW_PERCENTILE
    samples["manual_low_flow_high_pb_low_zinc"] = (
        samples["manual_low_flow"] & samples["high_pb_low_zinc"]
    )
    samples["manual_low_flow_high_pb_zn_ratio"] = (
        samples["manual_low_flow"] & samples["high_pb_zn_ratio"]
    )
    samples["manual_flow_class"] = np.select(
        [
            samples["manual_flow_percentile"].isna(),
            samples["manual_flow_percentile"] <= VERY_LOW_FLOW_PERCENTILE,
            samples["manual_flow_percentile"] <= LOW_FLOW_PERCENTILE,
        ],
        ["missing_flow", "very_low_flow", "low_flow"],
        default="normal_or_high_flow",
    )

    samples["automatic_has_flow_match"] = samples["has_flow_match"].fillna(False).astype(bool)
    both = samples["automatic_has_flow_match"] & samples["manual_has_flow_match"]
    samples["both_methods_have_flow"] = both
    samples["low_flow_status_changed"] = False
    samples.loc[both, "low_flow_status_changed"] = (
        samples.loc[both, "low_flow"].astype(bool)
        != samples.loc[both, "manual_low_flow"].astype(bool)
    )
    samples["flow_comparison_category"] = np.select(
        [
            both & samples["low_flow"].astype(bool) & samples["manual_low_flow"],
            both & ~samples["low_flow"].astype(bool) & ~samples["manual_low_flow"],
            both & ~samples["low_flow"].astype(bool) & samples["manual_low_flow"],
            both & samples["low_flow"].astype(bool) & ~samples["manual_low_flow"],
            samples["automatic_has_flow_match"] & ~samples["manual_has_flow_match"],
            ~samples["automatic_has_flow_match"] & samples["manual_has_flow_match"],
        ],
        [
            "Low flow in both",
            "Not low flow in both",
            "Became low flow",
            "No longer low flow",
            "Manual flow unavailable",
            "Manual flow added",
        ],
        default="Flow unavailable in both",
    )
    samples["high_pbzn_event_changed"] = False
    samples.loc[both, "high_pbzn_event_changed"] = (
        samples.loc[both, "low_flow_high_pb_zn_ratio"].fillna(False).astype(bool)
        != samples.loc[both, "manual_low_flow_high_pb_zn_ratio"].fillna(False).astype(bool)
    )
    samples.to_csv(PROCESSED / "pbzn_manual_flow_validation_samples.csv", index=False)
    return samples, both


def station_summary(samples: pd.DataFrame, prefix: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for station_id, group in samples.groupby("station_id"):
        if prefix == "automatic":
            has_flow = group["automatic_has_flow_match"]
            low_flow = group["low_flow"].fillna(False).astype(bool)
            low_high_ratio = group["low_flow_high_pb_zn_ratio"].fillna(False).astype(bool)
            low_high_pb_low_zn = group["low_flow_high_pb_low_zinc"].fillna(False).astype(bool)
            flow_percentile = group["flow_percentile"]
        else:
            has_flow = group["manual_has_flow_match"]
            low_flow = group["manual_low_flow"].fillna(False).astype(bool)
            low_high_ratio = group["manual_low_flow_high_pb_zn_ratio"].fillna(False).astype(bool)
            low_high_pb_low_zn = group["manual_low_flow_high_pb_low_zinc"].fillna(False).astype(bool)
            flow_percentile = group["manual_flow_percentile"]
        with_flow = group[has_flow]
        n_flow = int(has_flow.sum())
        n_low_high_ratio = int(low_high_ratio[has_flow].sum())
        n_low_high_pb_low_zn = int(low_high_pb_low_zn[has_flow].sum())
        rows.append(
            {
                "station_id": station_id,
                f"{prefix}_n_samples": len(group),
                f"{prefix}_n_with_flow": n_flow,
                f"{prefix}_n_low_flow": int(low_flow[has_flow].sum()),
                f"{prefix}_n_low_flow_high_pbzn": n_low_high_ratio,
                f"{prefix}_n_low_flow_high_pb_low_zn": n_low_high_pb_low_zn,
                f"{prefix}_pct_low_flow_high_pbzn": n_low_high_ratio / n_flow if n_flow else np.nan,
                f"{prefix}_median_flow_percentile": flow_percentile[has_flow].median() if n_flow else np.nan,
            }
        )
    summary = pd.DataFrame(rows)
    summary[f"{prefix}_systematic_candidate"] = (
        (summary[f"{prefix}_n_with_flow"] >= 3)
        & (
            (summary[f"{prefix}_n_low_flow_high_pb_low_zn"] >= 2)
            | (summary[f"{prefix}_n_low_flow_high_pbzn"] >= 3)
            | (
                (summary[f"{prefix}_pct_low_flow_high_pbzn"] >= 0.30)
                & (summary[f"{prefix}_n_low_flow_high_pbzn"] >= 2)
            )
        )
    )
    return summary


def build_station_sensitivity(
    selected: pd.DataFrame, samples: pd.DataFrame
) -> pd.DataFrame:
    automatic = station_summary(samples, "automatic")
    manual = station_summary(samples, "manual")
    sensitivity = selected[
        [
            "station_id",
            "station_name",
            "selection_group",
            "automatic_distance_band",
            "nrfa_station_id",
            "nrfa_station_name",
            "manual_nrfa_id",
            "manual_nrfa_name",
            "manual_decision",
            "manual_confidence",
            "exact_station_agreement",
        ]
    ].merge(automatic, on="station_id", how="left")
    sensitivity = sensitivity.merge(manual, on="station_id", how="left")
    sample_changes = (
        samples.groupby("station_id", as_index=False)
        .agg(
            n_samples_with_flow_under_both=("both_methods_have_flow", "sum"),
            n_low_flow_status_changes=("low_flow_status_changed", "sum"),
            n_high_pbzn_event_changes=("high_pbzn_event_changed", "sum"),
        )
    )
    sensitivity = sensitivity.merge(sample_changes, on="station_id", how="left")
    sensitivity["low_flow_status_change_rate"] = (
        sensitivity["n_low_flow_status_changes"]
        / sensitivity["n_samples_with_flow_under_both"].replace(0, np.nan)
    )
    manual_evaluable = sensitivity["manual_n_with_flow"] >= 3
    sensitivity["candidate_status_change"] = np.select(
        [
            ~manual_evaluable,
            sensitivity["automatic_systematic_candidate"]
            & sensitivity["manual_systematic_candidate"],
            sensitivity["automatic_systematic_candidate"]
            & ~sensitivity["manual_systematic_candidate"],
            ~sensitivity["automatic_systematic_candidate"]
            & sensitivity["manual_systematic_candidate"],
        ],
        [
            "Not evaluable after manual match",
            "Candidate in both",
            "Candidate lost after manual match",
            "New candidate after manual match",
        ],
        default="Not candidate in either",
    )
    rank_columns = [
        "automatic_n_low_flow_high_pbzn",
        "manual_n_low_flow_high_pbzn",
    ]
    for column in rank_columns:
        rank_name = column.replace("_n_low_flow_high_pbzn", "_candidate_rank")
        sensitivity[rank_name] = np.nan
        prefix = column.split("_", 1)[0]
        rank_mask = (
            sensitivity["selection_group"].eq("Systematic candidate")
            & sensitivity[f"{prefix}_n_with_flow"].ge(3)
        )
        sensitivity.loc[rank_mask, rank_name] = sensitivity.loc[rank_mask, column].rank(
            method="min", ascending=False
        )
    sensitivity.to_csv(PROCESSED / "manual_matching_station_sensitivity.csv", index=False)
    return sensitivity


def build_metrics(
    selected: pd.DataFrame,
    samples: pd.DataFrame,
    sensitivity: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    both = samples["both_methods_have_flow"]
    changed = samples.loc[both, "low_flow_status_changed"]
    candidate_rows = sensitivity[sensitivity["selection_group"] == "Systematic candidate"]
    control_rows = sensitivity[sensitivity["selection_group"] == "Stratified control"]
    evaluable_candidates = candidate_rows[candidate_rows["manual_n_with_flow"] >= 3]
    unevaluable_candidates = candidate_rows[candidate_rows["manual_n_with_flow"] < 3]
    percentile_correlation = samples.loc[both, ["flow_percentile", "manual_flow_percentile"]].corr(
        method="spearman"
    ).iloc[0, 1]
    rank_correlation = candidate_rows[
        ["automatic_candidate_rank", "manual_candidate_rank"]
    ].corr(method="spearman").iloc[0, 1]

    metrics = [
        ("Validation sites", len(selected), "sites"),
        ("Systematic candidate sites", (selected["selection_group"] == "Systematic candidate").sum(), "sites"),
        ("Stratified control sites", (selected["selection_group"] == "Stratified control").sum(), "sites"),
        ("Exact automatic/manual station agreement", selected["exact_station_agreement"].sum(), "sites"),
        ("Exact station agreement rate", selected["exact_station_agreement"].mean(), "proportion"),
        ("Automatic station retained", selected["manual_decision"].eq("Retain automatic match").sum(), "sites"),
        ("Automatic station replaced", selected["manual_decision"].str.startswith("Replace").sum(), "sites"),
        ("Hydrological gauge without period overlap", selected["manual_decision"].str.contains("no sample-period overlap").sum(), "sites"),
        ("No suitable NRFA mean-flow gauge", selected["manual_decision"].eq("No suitable NRFA mean-flow gauge").sum(), "sites"),
        ("Selected Pb/Zn samples", len(samples), "samples"),
        ("Automatic same-day flow samples", samples["automatic_has_flow_match"].sum(), "samples"),
        ("Manual same-day flow samples", samples["manual_has_flow_match"].sum(), "samples"),
        ("Samples with flow under both methods", both.sum(), "samples"),
        ("Low-flow status changes among paired methods", changed.sum(), "samples"),
        ("Low-flow status change rate", changed.mean() if len(changed) else np.nan, "proportion"),
        ("Pb/Zn high-ratio event classification changes among paired methods", samples.loc[both, "high_pbzn_event_changed"].sum(), "samples"),
        ("Automatic/manual flow-percentile Spearman correlation", percentile_correlation, "rho"),
        ("Original candidate sites retained after manual matching", evaluable_candidates["manual_systematic_candidate"].sum(), "sites"),
        ("Original candidate sites lost after manual matching", (~evaluable_candidates["manual_systematic_candidate"]).sum(), "sites"),
        ("Original candidate sites not evaluable after manual matching", len(unevaluable_candidates), "sites"),
        ("New candidates among nine controls", control_rows["manual_systematic_candidate"].sum(), "sites"),
        ("Candidate-ranking Spearman correlation", rank_correlation, "rho"),
    ]
    metrics_df = pd.DataFrame(metrics, columns=["metric", "value", "unit"])

    by_distance = (
        selected.groupby("automatic_distance_band", observed=False)
        .agg(
            n_sites=("station_id", "size"),
            n_exact_agreement=("exact_station_agreement", "sum"),
        )
        .reset_index()
    )
    order = ["Short (<2 km)", "Medium (2-<7 km)", "Long (>=7 km)"]
    by_distance["automatic_distance_band"] = pd.Categorical(
        by_distance["automatic_distance_band"], categories=order, ordered=True
    )
    by_distance = by_distance.sort_values("automatic_distance_band")
    by_distance["agreement_rate"] = by_distance["n_exact_agreement"] / by_distance["n_sites"]

    metrics_df.to_csv(PROCESSED / "manual_matching_validation_metrics.csv", index=False)
    by_distance.to_csv(PROCESSED / "manual_matching_agreement_by_distance.csv", index=False)
    return metrics_df, by_distance


def style_excel(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column[:200]]
            width = min(max(max((len(value) for value in values), default=0) + 2, 11), 48)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_excel_outputs(
    selected: pd.DataFrame,
    samples: pd.DataFrame,
    sensitivity: pd.DataFrame,
    metrics: pd.DataFrame,
    by_distance: pd.DataFrame,
) -> None:
    site_columns = [
        "selection_group",
        "station_id",
        "station_name",
        "station_type",
        "wfd_c2_mgt_catchment_name",
        "easting",
        "northing",
        "n_pbzn_samples",
        "n_with_flow",
        "automatic_distance_band",
        "nrfa_station_id",
        "nrfa_station_name",
        "nrfa_river",
        "flow_match_distance_km",
        "hydrologic_best_nrfa_id",
        "hydrologic_best_nrfa_name",
        "manual_nrfa_id",
        "manual_nrfa_name",
        "manual_nrfa_river",
        "manual_match_distance_km",
        "manual_decision",
        "automatic_match_assessment",
        "exact_station_agreement",
        "manual_confidence",
        "hydrological_basis",
        "evidence_title",
        "evidence_url",
        "nrfa_station_url",
        "manual_gdf_start_date",
        "manual_gdf_end_date",
    ]
    sources_df = pd.DataFrame(
        [{"source_key": key, **value} for key, value in SOURCES.items()]
    )
    definitions = pd.DataFrame(
        [
            ["Systematic candidate", "All 21 stations flagged by the automatic 7.3 analysis."],
            ["Stratified control", "Nine non-candidate stations: three short, three medium and three long automatic match distances, spanning several catchments."],
            ["Exact station agreement", "Automatic and manual methods select the same NRFA station."],
            ["Low flow", "Daily flow percentile <=0.25 (the 25th percentile) for the selected NRFA station's full available daily record."],
            ["Manual match", "Desk-based hydrological judgement using receiving-water evidence, station names, coordinates, NRFA metadata and period overlap."],
            ["Important limitation", "This is a targeted 30-site validation, not a manual rematch of all 438 Pb/Zn stations."],
        ],
        columns=["term", "definition"],
    )

    manual_path = RESULTS / "7.4_Manual_Matching.xlsx"
    with pd.ExcelWriter(manual_path, engine="openpyxl") as writer:
        selected[site_columns].to_excel(writer, sheet_name="30_Sites", index=False)
        selected.loc[selected["selection_group"] == "Systematic candidate", site_columns].to_excel(
            writer, sheet_name="21_Candidates", index=False
        )
        selected.loc[selected["selection_group"] == "Stratified control", site_columns].to_excel(
            writer, sheet_name="9_Controls", index=False
        )
        by_distance.to_excel(writer, sheet_name="Agreement_By_Distance", index=False)
        sources_df.to_excel(writer, sheet_name="Sources", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    style_excel(manual_path)

    sample_columns = [
        "station_id",
        "station_name",
        "selection_group",
        "sample_day",
        "lead",
        "zinc",
        "pb_zn_ratio",
        "high_pb_zn_ratio",
        "nrfa_station_id",
        "nrfa_station_name",
        "flow_m3s",
        "flow_percentile",
        "low_flow",
        "manual_nrfa_id",
        "manual_nrfa_name",
        "manual_flow_m3s",
        "manual_flow_percentile",
        "manual_low_flow",
        "flow_comparison_category",
        "low_flow_status_changed",
        "high_pbzn_event_changed",
    ]
    ranking = sensitivity.sort_values(
        ["manual_systematic_candidate", "manual_n_low_flow_high_pbzn", "manual_pct_low_flow_high_pbzn"],
        ascending=[False, False, False],
    )
    comparison_path = RESULTS / "7.4_Matching_Comparison.xlsx"
    with pd.ExcelWriter(comparison_path, engine="openpyxl") as writer:
        metrics.to_excel(writer, sheet_name="Summary_Metrics", index=False)
        by_distance.to_excel(writer, sheet_name="Agreement_By_Distance", index=False)
        sensitivity.to_excel(writer, sheet_name="Station_Sensitivity", index=False)
        ranking.to_excel(writer, sheet_name="Manual_Ranking", index=False)
        samples[sample_columns].to_excel(writer, sheet_name="Sample_Comparison", index=False)
    style_excel(comparison_path)


def make_figures(
    selected: pd.DataFrame,
    samples: pd.DataFrame,
    sensitivity: pd.DataFrame,
    by_distance: pd.DataFrame,
) -> None:
    plt.style.use("seaborn-v0_8-whitegrid")
    colors = {True: "#2D6A4F", False: "#C56B3D"}

    fig, ax = plt.subplots(figsize=(8.6, 5.4))
    x = np.arange(len(by_distance))
    agreed = by_distance["n_exact_agreement"].to_numpy()
    changed = by_distance["n_sites"].to_numpy() - agreed
    ax.bar(x, agreed, color=colors[True], label="Same NRFA station")
    ax.bar(x, changed, bottom=agreed, color=colors[False], label="Changed or unavailable")
    ax.set_xticks(x, by_distance["automatic_distance_band"].astype(str))
    ax.set_ylabel("Validation sites")
    ax.set_title("Automatic versus manual station agreement by distance")
    for idx, row in by_distance.reset_index(drop=True).iterrows():
        ax.text(idx, row["n_sites"] + 0.2, f"{row['agreement_rate']:.0%}", ha="center", fontweight="bold")
    ax.legend(frameon=False, ncol=2, loc="upper right")
    fig.tight_layout()
    fig.savefig(RESULTS / "7.4_Matching_Agreement.png", dpi=220)
    plt.close(fig)

    category_order = [
        "Low flow in both",
        "Not low flow in both",
        "Became low flow",
        "No longer low flow",
        "Manual flow unavailable",
        "Manual flow added",
        "Flow unavailable in both",
    ]
    category_counts = samples["flow_comparison_category"].value_counts().reindex(category_order, fill_value=0)
    fig, ax = plt.subplots(figsize=(9.2, 5.8))
    palette = ["#2D6A4F", "#5B7F8C", "#D6A84B", "#C56B3D", "#8D6E63", "#6A8EAE", "#B0B0B0"]
    bars = ax.barh(category_counts.index, category_counts.values, color=palette)
    ax.invert_yaxis()
    ax.set_xlabel("Pb/Zn station-date samples")
    ax.set_title("Effect of manual matching on low-flow classification")
    ax.bar_label(bars, padding=4)
    fig.tight_layout()
    fig.savefig(RESULTS / "7.4_Flow_Class_Changes.png", dpi=220)
    plt.close(fig)

    candidates = sensitivity[sensitivity["selection_group"] == "Systematic candidate"].copy()
    candidates["max_events"] = candidates[
        ["automatic_n_low_flow_high_pbzn", "manual_n_low_flow_high_pbzn"]
    ].max(axis=1)
    candidates = candidates.sort_values("max_events", ascending=False).head(12).sort_values("max_events")
    labels = candidates["station_name"].str.slice(0, 42)
    y = np.arange(len(candidates))
    fig, ax = plt.subplots(figsize=(10.5, 7.3))
    height = 0.36
    ax.barh(y - height / 2, candidates["automatic_n_low_flow_high_pbzn"], height, color="#5B7F8C", label="Automatic")
    ax.barh(y + height / 2, candidates["manual_n_low_flow_high_pbzn"], height, color="#C56B3D", label="Manual")
    ax.set_yticks(y, labels)
    ax.set_xlabel("Low-flow samples with high Pb/Zn")
    ax.set_title("Candidate-site ranking sensitivity to flow-station choice")
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(RESULTS / "7.4_Candidate_Ranking_Comparison.png", dpi=220)
    plt.close(fig)


def xml_text(value: object) -> str:
    return escape(str(value))


def paragraph(text: str = "", style: str = "Normal", bold: bool = False) -> str:
    bold_xml = "<w:rPr><w:b/></w:rPr>" if bold else ""
    return (
        f'<w:p><w:pPr><w:pStyle w:val="{style}"/></w:pPr>'
        f'<w:r>{bold_xml}<w:t xml:space="preserve">{xml_text(text)}</w:t></w:r></w:p>'
    )


def bullet(text: str) -> str:
    return (
        '<w:p><w:pPr><w:pStyle w:val="ListParagraph"/>'
        '<w:numPr><w:ilvl w:val="0"/><w:numId w:val="1"/></w:numPr></w:pPr>'
        f'<w:r><w:t xml:space="preserve">{xml_text(text)}</w:t></w:r></w:p>'
    )


def docx_table(headers: list[str], rows: list[list[object]]) -> str:
    def cell(value: object, header: bool = False) -> str:
        bold = "<w:rPr><w:b/></w:rPr>" if header else ""
        return (
            '<w:tc><w:tcPr><w:tcW w:w="2200" w:type="dxa"/></w:tcPr>'
            f'<w:p><w:r>{bold}<w:t>{xml_text(value)}</w:t></w:r></w:p></w:tc>'
        )

    table_parts = [
        '<w:tbl><w:tblPr><w:tblStyle w:val="TableGrid"/><w:tblW w:w="0" w:type="auto"/></w:tblPr>',
        "<w:tr>" + "".join(cell(value, True) for value in headers) + "</w:tr>",
    ]
    for row in rows:
        table_parts.append("<w:tr>" + "".join(cell(value) for value in row) + "</w:tr>")
    table_parts.append("</w:tbl>")
    return "".join(table_parts)


def make_docx(path: Path, body_parts: list[str]) -> None:
    document_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>{''.join(body_parts)}<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>
<w:pgMar w:top="1134" w:right="1134" w:bottom="1134" w:left="1134"/></w:sectPr></w:body></w:document>'''
    styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/>
<w:pPr><w:spacing w:after="120" w:line="300" w:lineRule="auto"/></w:pPr>
<w:rPr><w:rFonts w:ascii="Aptos" w:hAnsi="Aptos" w:eastAsia="Microsoft YaHei"/><w:sz w:val="22"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/>
<w:rPr><w:rFonts w:ascii="Aptos Display" w:hAnsi="Aptos Display" w:eastAsia="Microsoft YaHei"/><w:b/><w:color w:val="1F4E5F"/><w:sz w:val="36"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="Heading 1"/><w:basedOn w:val="Normal"/>
<w:pPr><w:keepNext/><w:spacing w:before="280" w:after="100"/></w:pPr>
<w:rPr><w:rFonts w:ascii="Aptos Display" w:hAnsi="Aptos Display" w:eastAsia="Microsoft YaHei"/><w:b/><w:color w:val="1F4E5F"/><w:sz w:val="28"/></w:rPr></w:style>
<w:style w:type="paragraph" w:styleId="ListParagraph"><w:name w:val="List Paragraph"/><w:basedOn w:val="Normal"/>
<w:pPr><w:ind w:left="720" w:hanging="360"/><w:spacing w:after="80"/></w:pPr></w:style>
<w:style w:type="table" w:styleId="TableGrid"><w:name w:val="Table Grid"/><w:tblPr><w:tblBorders>
<w:top w:val="single" w:sz="4"/><w:left w:val="single" w:sz="4"/><w:bottom w:val="single" w:sz="4"/>
<w:right w:val="single" w:sz="4"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/>
</w:tblBorders></w:tblPr></w:style></w:styles>'''
    numbering_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:numbering xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:abstractNum w:abstractNumId="0"><w:lvl w:ilvl="0"><w:start w:val="1"/><w:numFmt w:val="bullet"/><w:lvlText w:val="•"/>
<w:pPr><w:ind w:left="720" w:hanging="360"/></w:pPr></w:lvl></w:abstractNum>
<w:num w:numId="1"><w:abstractNumId w:val="0"/></w:num></w:numbering>'''
    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
<Override PartName="/word/numbering.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.numbering+xml"/>
</Types>'''
    root_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>'''
    doc_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/numbering" Target="numbering.xml"/>
</Relationships>'''
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("word/document.xml", document_xml)
        archive.writestr("word/styles.xml", styles_xml)
        archive.writestr("word/numbering.xml", numbering_xml)
        archive.writestr("word/_rels/document.xml.rels", doc_rels)


def metric_value(metrics: pd.DataFrame, name: str) -> float:
    return float(metrics.loc[metrics["metric"] == name, "value"].iloc[0])


def write_word_summaries(
    selected: pd.DataFrame,
    metrics: pd.DataFrame,
    by_distance: pd.DataFrame,
) -> None:
    agreement_rate = metric_value(metrics, "Exact station agreement rate")
    low_change_rate = metric_value(metrics, "Low-flow status change rate")
    both_samples = int(metric_value(metrics, "Samples with flow under both methods"))
    low_change_count = int(metric_value(metrics, "Low-flow status changes among paired methods"))
    event_change_count = int(
        metric_value(metrics, "Pb/Zn high-ratio event classification changes among paired methods")
    )
    candidate_retained = int(metric_value(metrics, "Original candidate sites retained after manual matching"))
    candidate_lost = int(metric_value(metrics, "Original candidate sites lost after manual matching"))
    candidate_not_evaluable = int(
        metric_value(metrics, "Original candidate sites not evaluable after manual matching")
    )
    rank_correlation = metric_value(metrics, "Candidate-ranking Spearman correlation")
    replacements = int(metric_value(metrics, "Automatic station replaced"))
    no_gauge = int(metric_value(metrics, "No suitable NRFA mean-flow gauge"))
    no_overlap = int(metric_value(metrics, "Hydrological gauge without period overlap"))

    distance_rows = [
        [
            row["automatic_distance_band"],
            int(row["n_sites"]),
            int(row["n_exact_agreement"]),
            f"{row['agreement_rate']:.1%}",
        ]
        for _, row in by_distance.iterrows()
    ]
    distance_label_cn = {
        "Short (<2 km)": "短距离（<2 km）",
        "Medium (2-<7 km)": "中距离（2至<7 km）",
        "Long (>=7 km)": "长距离（>=7 km）",
    }
    distance_rows_cn = [
        [distance_label_cn.get(row[0], row[0]), row[1], row[2], row[3]]
        for row in distance_rows
    ]
    recommendation = (
        "Use a hybrid catchment-aware approach: retain the nearest-station match for sites below 2 km only after a same-river check, "
        "and manually verify or correct all matches at 2 km or more. Distance alone should not be treated as proof of hydrological connectivity."
    )

    english_body = [
        paragraph("Manual Flow-Station Matching Validation", "Title"),
        paragraph("Section 7.4 / Project Methods 3.9 and Results 4.11"),
        paragraph("Scope", "Heading1"),
        paragraph("The validation covered all 21 systematic candidate sites from the automatic Pb/Zn-flow analysis and nine stratified non-candidate controls across short, medium and long automatic match distances."),
        paragraph("Method", "Heading1"),
        bullet("The automatic nearest-NRFA result was retained as the baseline."),
        bullet("Each selected water-quality site was checked against receiving-water evidence, station naming, coordinates, NRFA river metadata and period of record."),
        bullet("The manual station was used to retrieve same-day daily mean flow and to recalculate station-specific flow percentiles using the same <=25th-percentile low-flow definition."),
        paragraph("Station agreement", "Heading1"),
        paragraph(f"The automatic and manual methods selected the same NRFA station for {int(selected['exact_station_agreement'].sum())} of 30 sites ({agreement_rate:.1%}). {replacements} sites required replacement, {no_overlap} had a hydrologically relevant station without sample-period overlap, and {no_gauge} had no suitable NRFA mean-flow gauge."),
        docx_table(["Automatic distance", "Sites", "Same station", "Agreement"], distance_rows),
        paragraph("Sensitivity of Pb/Zn-flow results", "Heading1"),
        paragraph(
            f"There were {both_samples:,} station-date samples with daily flow under both methods. "
            f"Low-flow status changed for {low_change_count:,} samples ({low_change_rate:.1%}), and {event_change_count:,} low-flow/high-Pb/Zn event classifications changed. "
            f"Of the original 21 systematic candidates, {candidate_retained} remained candidates, {candidate_lost} no longer met the original rule, "
            f"and {candidate_not_evaluable} could not be re-evaluated because the hydrologically relevant gauge did not overlap the sampling period. "
            f"Candidate ranking remained related but not identical (Spearman rho={rank_correlation:.2f})."
        ),
        paragraph("Conclusion", "Heading1"),
        paragraph(recommendation, bold=True),
        paragraph("Important limitation", "Heading1"),
        paragraph("This is a targeted desk-based validation of 30 sites. It does not manually rematch all 438 Pb/Zn stations. Some manual stations are downstream main-river proxies for small ungauged mine tributaries, and NRFA daily mean flow is not instantaneous flow at the sampling time. The site-level evidence and source links are recorded in 7.4_Manual_Matching.xlsx."),
    ]
    make_docx(RESULTS / "summary.docx", english_body)

    chinese_body = [
        paragraph("7.4 人工流量站匹配验证总结", "Title"),
        paragraph("对应项目 Methodology 3.9 与 Results 4.11"),
        paragraph("验证范围", "Heading1"),
        paragraph("本次验证包括自动分析识别出的全部21个系统性候选站点，以及按自动匹配距离分层选择的9个普通站点，共30个站点。"),
        paragraph("方法", "Heading1"),
        bullet("保留最近距离自动匹配作为基准。"),
        bullet("根据接收水体资料、站点名称和坐标、NRFA河流信息、上下游关系及数据时期，对每个站点进行人工水文核对。"),
        bullet("使用人工选择的流量站重新匹配同日平均流量，并继续采用流量站自身第25百分位作为低流量阈值。"),
        paragraph("匹配一致性", "Heading1"),
        paragraph(f"30个站点中，自动方法与人工方法有{int(selected['exact_station_agreement'].sum())}个选择了同一NRFA站，一致率为{agreement_rate:.1%}。其中{replacements}个需要更换，{no_overlap}个存在水文上合适但与采样时期不重叠的站点，{no_gauge}个没有合适的NRFA平均流量站。"),
        docx_table(["自动匹配距离", "站点数", "相同流量站", "一致率"], distance_rows_cn),
        paragraph("对Pb/Zn–流量结果的影响", "Heading1"),
        paragraph(
            f"两种方法均有同日流量的站点日期样本为{both_samples:,}条，其中{low_change_count:,}条（{low_change_rate:.1%}）的低流量分类发生改变，"
            f"{event_change_count:,}条低流量高Pb/Zn事件分类发生改变。原来的21个系统性候选站点中，人工匹配后{candidate_retained}个仍为候选，"
            f"{candidate_lost}个不再满足原候选规则，另有{candidate_not_evaluable}个因合适流量站与采样时期不重叠而无法重新评估。"
            f"候选排名保持中等偏强相关，但并不完全一致（Spearman rho={rank_correlation:.2f}）。"
        ),
        paragraph("最终建议", "Heading1"),
        paragraph("采用经过验证的混合方法：自动距离小于2 km的匹配仍需进行同河流快速检查；距离达到或超过2 km的站点必须进行人工水文核对或修正。不能仅凭距离证明两个站点水文相连。", bold=True),
        paragraph("限制", "Heading1"),
        paragraph("这是针对30个站点的桌面人工验证，并未人工重配全部438个Pb/Zn站点。部分矿区小支流没有独立流量站，只能使用下游主河道代理站；NRFA日平均流量也不等于采样时刻的瞬时流量。各站点判断依据和资料链接保存在7.4_Manual_Matching.xlsx中。"),
    ]
    make_docx(RESULTS / "7.4_Summary_CN.docx", chinese_body)


def copy_project_outputs() -> None:
    copies = {
        RESULTS / "7.4_Manual_Matching.xlsx": PROJECT_RESULTS / "3.9_Manual_Flow_Matching.xlsx",
        RESULTS / "7.4_Matching_Comparison.xlsx": PROJECT_RESULTS / "4.11_Matching_Comparison.xlsx",
        RESULTS / "7.4_Matching_Agreement.png": PROJECT_RESULTS / "4.11_Matching_Agreement.png",
        RESULTS / "7.4_Flow_Class_Changes.png": PROJECT_RESULTS / "4.11_Flow_Class_Changes.png",
        RESULTS / "7.4_Candidate_Ranking_Comparison.png": PROJECT_RESULTS / "4.11_Candidate_Ranking_Comparison.png",
        RESULTS / "summary.docx": PROJECT_RESULTS / "4.11_Summary.docx",
        RESULTS / "7.4_Summary_CN.docx": PROJECT_RESULTS / "4.11_Summary_CN.docx",
    }
    for source, destination in copies.items():
        shutil.copy2(source, destination)


def make_package() -> None:
    package_path = RESULTS / "7.4_Manual_Matching_Validation_Package.zip"
    files = [
        "7.4_Manual_Matching.xlsx",
        "7.4_Matching_Comparison.xlsx",
        "7.4_Matching_Agreement.png",
        "7.4_Flow_Class_Changes.png",
        "7.4_Candidate_Ranking_Comparison.png",
        "summary.docx",
        "7.4_Summary_CN.docx",
    ]
    with ZipFile(package_path, "w", ZIP_DEFLATED) as archive:
        for filename in files:
            archive.write(RESULTS / filename, arcname=filename)


def main() -> None:
    ensure_directories()
    selected, _ = load_validation_sites()
    manual_station_ids = selected["manual_nrfa_id"].dropna().astype(int).unique().tolist()
    flows = load_manual_flows(manual_station_ids)
    samples, _ = build_sample_comparison(selected, flows)
    sensitivity = build_station_sensitivity(selected, samples)
    metrics, by_distance = build_metrics(selected, samples, sensitivity)
    write_excel_outputs(selected, samples, sensitivity, metrics, by_distance)
    make_figures(selected, samples, sensitivity, by_distance)
    write_word_summaries(selected, metrics, by_distance)
    copy_project_outputs()
    make_package()

    print(f"Validation sites: {len(selected)}")
    print(f"Exact station agreement: {int(selected['exact_station_agreement'].sum())}/30")
    print(f"Selected Pb/Zn samples: {len(samples):,}")
    print(f"Manual same-day flow samples: {int(samples['manual_has_flow_match'].sum()):,}")
    print(f"Wrote outputs to: {RESULTS}")


if __name__ == "__main__":
    main()
